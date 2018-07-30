'''
This is designed to run as part of ipchecker.py at the very end.
It takes the unsafe IP's and runs them through IBM X-Force Exchange for the reason for blacklisting
If no category returns, the system defaults to a 'Potential False Positive' response.
Human logic is needed to justify if too few blacklists are available to consider it a false positive
'''

import openpyxl
import requests
from requests.auth import HTTPBasicAuth #needed to authenticate with the API key

wb = openpyxl.load_workbook('example_testfile.xlsx') #fill in with whatever input you want for the information
ws = wb['Sheet1']

value = 2
letter = 'A'

ip = 'i'

print_value = str(value)
cell = letter+print_value #starts us at cell A2

while ws[cell].value: #while the cell hodls value
	print_value = str(value)
	cell = 'B'+print_value #looks at whether it's safe or not

	if ws[cell].value == 'No': #if not safe
		cell = letter+print_value
		ip = str(ws[cell].value)
		print('Checking IBM X-Force Exchange for: '+ip)#just a line of code to help us be aware that the program is still running
		response = requests.get("https://api.xforce.ibmcloud.com/ipr/history/"+ip, auth=HTTPBasicAuth('f0d4525f-a4da-42ef-ba45-17d82d812531', '92209c34-896d-4f08-b787-723f2b71b336'))
		response_data = response.json()
		#print(response_data) #If you ever need to problem test
		cell = 'D'+print_value
		ws[cell].value = ''.join(list(response_data['history'][-1]['categoryDescriptions'].keys()))
		if ws[cell].value == '': #if IBM reports back no errors with the site
			ws[cell].value = 'Potential False Positive'
	value += 1
	cell = letter+print_value
	#print("end of loop") #used for error testing
	wb.save('example_testfile.xlsx')
wb.save('example_testfile.xlsx')
