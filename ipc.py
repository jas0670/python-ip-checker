import re
import grequests
import csv
import requests
import time
import os
from openpyxl import *

class IPC():
	def __init__(self):
		self.startrow = 2
		self.filename = ''
		self.workbook = ''
		self.sheet = ''

	def load(self):
		# Start by getting Excel filename and opening it, get first sheet
		self.filename = input("Enter excel filename with extension (xls, xlsx): ")
		self.workbook = load_workbook(self.filename) #,on_demand = True)
		self.sheet = self.workbook.active

		# Menu options
		choice = 0
		while choice < 1 or choice > 3:
			print("\n*** Wipro Python IP Checker ***\n")
			print("		1) Start from beginning")
			print("		2) Start from last checked" )
			print("		3) Exit\n")
			choice = int(input("  > "))

		# Set starting row based off user choice
		self.startrow = 2
		if choice == 2:
			for x in range(2, self.sheet.max_row + 1):
				print(self.sheet.cell(x, 2))
				if self.sheet.cell(x, 2).value is None:
					self.startrow = x
					break
		elif choice == 3:
			exit()

	# Pulls data from the Xforce Exchange API
	def xforce(self):
		# Collect list of all Xforce Exchange API URLs for each IP
		# Due to restrictions of Xforce Exchange API, only request 10 IPs at a time
		# Code will gather all urls in one list, and then split into chunks for sending 10 at a tim
		responses = []
		urls = []
		for row in self.sheet.iter_rows(min_row=self.startrow, max_col=1, max_row=self.sheet.max_row):
			for cell in row:
				urls.append("https://api.xforce.ibmcloud.com/ipr/history/{}".format(cell.value))

		# ASYNC REQUESTS
		# This mostly works, and is faster, but occassionally drops requests
		#url_lists = [urls[x:x+10] for x in range(0, len(urls), 10)]
		#for l in url_lists:
		#	request_list = (grequests.get(u, timeout=1000) for u in l)
		#	print("Sending request of {} IPs...".format(len(l)))
		#	responses.extend(grequests.map(request_list))

		# SYNC BLOCKING REQUESTS
		for u in urls:
			print("Sending request for {}".format(u))
			# Key located below
			response = requests.get(u, auth=('f0d4525f-a4da-42ef-ba45-17d82d812531','92209c34-896d-4f08-b787-723f2b71b336'))
			responses.append(response)

		# Get the data from each response. Data currently includes score, and a comma separated
		# list of all categories that gave that score
		data = []
		for r in responses:
			try:
				current_data = r.json()
				data.append({'score': current_data['history'][-1]['score'], 'category': ", ".join(list(current_data['history'][-1]['categoryDescriptions'].keys()))})
			except:
				data.append({'score': "Timeout", 'category': "Timeout"})

		# Add scores back to spreadsheet and save
		# for cell in [x for t in self.sheet['B{}'.format(self.startrow):'B{}'.format(self.sheet.max_row)] for x in t]:
		for row in self.sheet.iter_rows(min_row=self.startrow, min_col=1, max_col=4, max_row=self.sheet.max_row):
			current_row = list(row)
			current_row[1].value = data[0]['score']
			current_row[2].value = data[0]['category']
			# Set hyperlink to the API
			current_row[3].hyperlink = "https://exchange.xforce.ibmcloud.com/ip/{}".format(current_row[0].value)
			current_row[3].value = "https://exchange.xforce.ibmcloud.com/ip/{}".format(current_row[0].value)
			data.pop(0)
			#for cell in row:
				#cell.value = data.pop(0)

	# Pulls data from the IPSpamList daily top IPs 1000 CSV
	def ipspamlist(self):
		with requests.Session() as s:
			try:
				download = s.get('http://www.ipspamlist.com/public_feeds.csv')

				decoded_content = download.content.decode('utf-8')
				cr = csv.reader(decoded_content.splitlines(), delimiter=',')
				my_list = list(cr)
				# Remove header row
				my_list.pop(0) 
				for row in my_list:
					# Put code here for processing each IP in the list
					# Index 2 is the IP
					print(row[2])
			except:
				print("Error attempting to download http://www.ipspamlist.com/public_feeds.csv")

	# Pulls data from the TotalVirus API
	def totalvirus(self):
		# Put Code here
		return None
	
	# Save the workbook
	def save(self):
		self.workbook.save(self.filename)


checker = IPC()
checker.load()
checker.xforce()
# Commented out for testing purposes
#checker.ipspamlist()
checker.save()