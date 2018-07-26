import virustotal2 #virus total has their own library that we used for this script
import os #imported for future usage, once we want the script to run on its own without user interface
import time #needed because virus total only takes 4 queries every minute. As of right now, it takes the script about an hour to run (7/23/2018)
import openpyxl #to read frome excel files 


wb = openpyxl.load_workbook('example_testfile.xlsx') #change file name to whatever you want to open to
ws = wb['Sheet2']
ss = wb['Sheet4'] #if we ever want to add information to a second sheet, the optionality is there

value = 2
letter = 'A'
vt = virustotal2.VirusTotal2("1fd2648228dde4b02c2438110543db84030033c0e4d9addd12a993b1c126d347") #API key. DO NOT DELETE

n = 0
ip = 'i'
print_value = str(value) #unorthodox way of iterating through everything in the manner I wanted to.
cell = letter+print_value

while ws[cell].value: #basically, while the cell has an obtainable value
	while n < 4:
		print_value = str(value)
		cell = letter+print_value
		ip = str(ws[cell].value)
		#print ip
		if ws[cell].value is None: #if there is no obtainable information in the loop
			#print "Inside the loop!"
			break
		ip_report = vt.retrieve(ip, raw = True) #the raw=True part is ESSENTIAL, otherwise the information retrieved is inaccessible.
		new_cell = 'B'+print_value
		ws[new_cell].value = ip_report
		n += 1
		value += 1
		wb.save('example_testfile.xlsx') #change file name to wherever you want the output

	#print n
	time.sleep(60) #makes the program pause for a minute to prevent the IP addresses from not timing out.
	n = 0


wb.save('example_testfile.xlsx') #change the final file name output to whatever you want.
