'''
This requires python3, make sure you have pip3 installed as well
Need to have installed from pip:
- openpyxl

NOTE: you cannot have your excel files open while this runs or the output to each file will not be saved.
'''

import os,sys #needed to run  bash scripts
import openpyxl #needed to mess with excel
import time #needed to make a pause
import subprocess #needed to make the program fun a shell script within the python environment
import re #regular expressions

wb = openpyxl.load_workbook('testfile.xlsx') #change file name to whatever
ws = wb['Sheet1']
ss = wb['Sheet2']

value = 2
letter = 'A'

ip = 'i'

print_value = str(value)
cell = letter+print_value
new_cell = letter+print_value
second_value = value #this is for the second sheet
x = 0
while x < 5:
	while ws[cell].value: #basically, while the cell has an obtainable value, do this
		print_value = str(value) #needed so the workbook can grab the appropriate cell 
		cell = letter+print_value #the cell in question
		ip = str(ws[cell].value) #the ip in question
		print('Checking for blacklistings for '+ip) #so we can keep track of what IP's are going through the system,. Mainly for testing purposes
		#if ws[cell].value is None: #if there is no obtainable information
		#	break
		ip_report = ('./blcheck ' + ip) #holds the command block to be executed
		child = subprocess.Popen(ip_report,shell=True,stdout=subprocess.PIPE) #execution of said commmand and the gathering of its output
		output = str(child.communicate()[0])
		output = re.findall(r'\d+',output)
		response = (int(output[-1]))
		if response > 0:
			#this following block takes the information and puts it into the second sheet of the excel document
			cell = 'B'+print_value # makes a new cell for the home sheet
			ws[cell].value = 'No' #stores that the ip address is not safe
			cell = 'C'+print_value #moves to the next column
			ws[cell].value = response #puts the number of blacklists in the home sheet
		else:
			#this is if the desired IP is safe and has no blacklists
			cell = 'B'+print_value
			ws[cell].value = 'Yes'
			cell = 'C'+print_value
			ws[cell].value = response
		value += 1 #goes to the next spot in the excel sheet
		print_value = str(value)
		cell = letter+print_value
		wb.save('example_testfile.xlsx')
	x += 1

subprocess.call("python3 ipcopy.py", shell=True)
